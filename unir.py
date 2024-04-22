import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
import tempfile
import shutil
import os

def save_uploaded_file(uploaded_file):
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx', mode='wb') as tmp_file:
            shutil.copyfileobj(uploaded_file, tmp_file)
            return tmp_file.name
    finally:
        uploaded_file.seek(0)  # Reset file pointer to the beginning for any subsequent operations

def compare_excel_files(file1, file2):
    path1 = save_uploaded_file(file1)
    path2 = save_uploaded_file(file2)
    try:
        df1 = pd.read_excel(path1, engine='openpyxl')
        df2 = pd.read_excel(path2, engine='openpyxl')

        if not (df1.shape == df2.shape and df1.columns.equals(df2.columns)):
            st.error("Las dimensiones o los índices de las columnas no coinciden en ambos archivos.")
            return None, None

        wb = openpyxl.Workbook()
        ws = wb.active
        fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        for col_num, value in enumerate(df1.columns, 1):
            ws.cell(row=1, column=col_num, value=value)

        row_offset = 2
        any_difference = False
        for row in range(df1.shape[0]):
            differences_in_row = False
            for col in range(df1.shape[1]):
                original_value = df1.iat[row, col]
                new_value = df2.iat[row, col]
                if pd.isna(original_value) and pd.isna(new_value):
                    continue  # Skip comparison for double blanks
                elif pd.isna(original_value) or pd.isna(new_value) or original_value != new_value:
                    cell = ws.cell(row=row_offset, column=col+1, value=new_value)
                    cell.fill = fill
                    differences_in_row = True
                else:
                    ws.cell(row=row_offset, column=col+1, value=original_value)

            if differences_in_row:
                any_difference = True
                row_offset += 1

        if any_difference:
            diff_file_path = 'differences.xlsx'
            wb.save(diff_file_path)
            return diff_file_path, pd.read_excel(diff_file_path, engine='openpyxl')
        else:
            st.info('No se encontraron diferencias.')
            return None, None
    finally:
        os.remove(path1)  # Clean up the temporary files
        os.remove(path2)

def main():
    st.title("Comparador de Archivos Excel")
    file1 = st.file_uploader("Sube el archivo Excel original", type=['xlsx'])
    file2 = st.file_uploader("Sube el archivo Excel a comparar", type=['xlsx'])
    
    if st.button("Comparar"):
        if file1 and file2:
            with st.spinner('Comparando archivos...'):
                result_file, diff = compare_excel_files(file1, file2)
                if result_file:
                    st.success('Comparación completada!')
                    st.download_button('Descargar archivo con diferencias', open(result_file, 'rb'), file_name=result_file)
                    st.dataframe(diff)
                else:
                    st.error('Error al comparar los archivos')

if __name__ == "__main__":
    main()
